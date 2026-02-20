import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Configuration
INPUT_FILE = 'data/sp800-53r4-to-r5-comparison-workbook.xlsx'
OUTPUT_FILE = 'content/nist_rev5_to_nist_rev4_crosswalk.csv'
RELATIONSHIPS_FILE = 'data/sp800-53r4-to-r5-comparison-relationships.xlsx'
SUMMARY_FILE = 'data/sp800-53r4-to-r5-comparison-summary.md'

# Classification constants (from nist_relationships.py)
ADDS            = ["adds control text", "adds parameter"]
REMOVES         = ["removes parameter", "removes control text"]
CHANGES_CONTROL = ["changes control text", "changes parameter"]
NEUTRAL         = ["changes discussion", "adds discussion", "changes title", "adds to", "n"]
NEW             = ["new base control", "new control enhancement"]

# Cell colours for each relationship type
COLORS = {
    "equal-to":        "C6EFCE",   # green
    "equivalent-to":   "FFEB9C",   # yellow
    "subset-of":       "BDD7EE",   # blue
    "superset-of":     "FCE4D6",   # orange
    "intersects-with": "E2EFDA",   # light green
    "no-relationship": "F2DCDB",   # pink/red
    "withdrawn":       "808080",   # dark grey (withdrawn in both Rev4 and Rev5)
    "withdrawn4":      "D9D9D9",   # grey (withdrawn in Rev4, not in Rev5)
    "restored5":       "E2CFDD",   # purple-ish (withdrawn in Rev4, restored in Rev5)
    "withdrawn5":      "C9C9C9",   # light grey (active in Rev4, withdrawn in Rev5)
    "withdrawn-error": "FF0000",   # red (unexpected withdrawn combination - needs review)
}

# Column structure for the crosswalk CSV (following the template pattern)
column_names = [
    '$$Source_Resource', 
    '$$Target_Resource', 
    '$$Map_Source_ID_Ref_list', 
    '$$Map_Target_ID_Ref_list', 
    '$$Map_Relationship', 
    '$Map_Confidence_Score', 
    '$Map_Coverage'
]

column_descriptions = [
    'A reference to a resource that has the source controls of a mapping.',
    'A reference to a resource that has the target controls of a mapping.',
    'A list of source reference IDs.',
    'A list of target reference IDs.',
    'The relationship type for the mapping entry.',
    'An estimation of the confidence that this mapping is correct and accurate expressed as percentage.',
    'An estimation of the percentage coverage of the targets by the sources.'
]

def transform_rev5_id(control_id):
    """
    Transform Rev 5 control IDs to OSCAL catalog format.
    Examples: 'AC-1' -> 'ac-1', 'AC-2(1)' -> 'ac-2.1'
    """
    if pd.isna(control_id):
        return ""
    
    # Convert to string and strip whitespace
    control_id = str(control_id).strip()
    
    # Convert to lowercase
    control_id = control_id.lower()
    
    # Remove leading zeros from numbers (AC-01 -> ac-1)
    # Split on hyphen, process the numeric part
    parts = control_id.split('-')
    if len(parts) == 2:
        family = parts[0]
        number_part = parts[1]
        
        # Handle enhancements in parentheses: ac-2(1) -> ac-2.1
        if '(' in number_part:
            base, enhancement = number_part.split('(')
            enhancement = enhancement.rstrip(')')
            # Remove leading zeros
            base = str(int(base)) if base.isdigit() else base
            enhancement = str(int(enhancement)) if enhancement.isdigit() else enhancement
            control_id = f"{family}-{base}.{enhancement}"
        else:
            # Just remove leading zeros from the number
            number_part = str(int(number_part)) if number_part.isdigit() else number_part
            control_id = f"{family}-{number_part}"
    
    return control_id

def transform_rev4_id(sort_as_id):
    """
    Transform Rev 4 SORT-AS format to OSCAL catalog format.
    Examples: 'AC-01-00' -> 'ac-1', 'AC-02-01' -> 'ac-2.1'
    """
    if pd.isna(sort_as_id):
        return ""
    
    # Convert to string and strip whitespace
    sort_as_id = str(sort_as_id).strip()
    
    # Convert to lowercase
    sort_as_id = sort_as_id.lower()
    
    # Split on hyphen: AC-01-00 -> ['ac', '01', '00']
    parts = sort_as_id.split('-')
    if len(parts) == 3:
        family = parts[0]
        base_num = parts[1]
        enhancement_num = parts[2]
        
        # Remove leading zeros
        base_num = str(int(base_num)) if base_num.isdigit() else base_num
        
        # If enhancement is 00, it's a base control
        if enhancement_num == '00':
            control_id = f"{family}-{base_num}"
        else:
            # It's an enhancement
            enhancement_num = str(int(enhancement_num)) if enhancement_num.isdigit() else enhancement_num
            control_id = f"{family}-{base_num}.{enhancement_num}"
    else:
        # Fallback: just return as-is
        control_id = sort_as_id
    
    return control_id

def classify_relationship(changed_elements, change_details):
    """
    Classify the OSCAL relationship between Rev 5 and Rev 4 controls.
    Based on logic from nist_relationships.py.
    
    Returns one of: equal-to, equivalent-to, superset-of, subset-of,
                    intersects-with, no-relationship, withdrawn, withdrawn4, withdrawn5
    """
    ce = str(changed_elements).strip().lower() if pd.notna(changed_elements) else ""
    cd = str(change_details).strip().lower() if pd.notna(change_details) else ""
    
    # Check for withdrawn combinations
    has_withdrawn_rev4 = "withdrawn in rev4" in cd
    has_restored_rev5 = "restored in rev5" in cd
    has_previously_withdrawn_rev4 = "previously withdrawn in rev4" in cd
    is_withdrawn_rev5 = ce == "withdrawn"
    
    # restored5: withdrawn in Rev4, explicitly restored in Rev5
    if has_withdrawn_rev4 and has_restored_rev5:
        return "restored5"
    
    # withdrawn4: previously withdrawn in Rev4, not in Rev5
    if has_previously_withdrawn_rev4:
        return "withdrawn4"
    
    # withdrawn: withdrawn in BOTH Rev4 and Rev5 (should not exist based on data)
    if has_withdrawn_rev4 and is_withdrawn_rev5 and not has_restored_rev5:
        return "withdrawn"
    
    # withdrawn5: active in Rev4, withdrawn in Rev5
    if is_withdrawn_rev5:
        return "withdrawn5"
    
    # Error check: unexpected withdrawn combination
    if has_withdrawn_rev4 and not has_restored_rev5 and not has_previously_withdrawn_rev4:
        # This should not happen - flag for review
        return "withdrawn-error"
    
    # New controls introduced in Rev5 with no Rev4 counterpart
    if any(n in ce for n in NEW):
        return "no-relationship"
    
    # Explicitly unchanged
    if ce == "n":
        return "equal-to"
    
    # Strip neutral-only lines to check for substantive changes
    lines = [l.strip().lower() for l in ce.split("\n") if l.strip()]
    substantive = [l for l in lines if not any(l.startswith(n) for n in NEUTRAL)]
    
    if not substantive:
        return "equivalent-to"
    
    has_adds = any(a in ce for a in ADDS)
    has_removes = any(r in ce for r in REMOVES)
    has_changes_control = any(c in ce for c in CHANGES_CONTROL)
    
    if has_changes_control:
        return "intersects-with"
    if has_adds and has_removes:
        return "intersects-with"
    if has_adds and not has_removes:
        return "superset-of"  # Rev5 gained requirements
    if has_removes and not has_adds:
        return "subset-of"  # Rev5 lost requirements
    
    return "intersects-with"

def generate_relationships_excel(df, input_file, output_file):
    """
    Generate an Excel file with OSCAL relationship classifications.
    Adds a new column to the existing workbook with color-coded relationships.
    """
    print(f"\nGenerating relationships Excel file...")
    
    # Load the workbook with openpyxl (preserves existing formatting)
    wb = load_workbook(input_file)
    ws = wb["Rev4 Rev5 Compared"]
    
    new_col = ws.max_column + 2   # leave one blank column as a gap
    
    # Header row (row 1)
    header = ws.cell(row=1, column=new_col, value="OSCAL Relationship\n(Rev5 → Rev4)")
    header.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header.fill = PatternFill("solid", fgColor="4472C4")
    header.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    
    # Sub-header row (row 2)
    sub = ws.cell(row=2, column=new_col, value="OSCAL Mapping (Rev5→Rev4)")
    sub.font = Font(bold=True, name="Arial", size=9)
    sub.fill = PatternFill("solid", fgColor="8EA9C1")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows start at Excel row 3
    for i, row in df.iterrows():
        excel_row = i + 3
        relationship = row["oscal_relationship"]
        cell = ws.cell(row=excel_row, column=new_col, value=relationship)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=COLORS.get(relationship, "FFFFFF"))
    
    ws.column_dimensions[get_column_letter(new_col)].width = 22
    
    # Save
    wb.save(output_file)
    print(f"✓ Created {output_file}")

def generate_summary_markdown(df, output_file, csv_stats=None):
    """
    Generate a markdown summary of the OSCAL relationships.
    
    Args:
        df: DataFrame with all controls and their relationships
        output_file: Path to output markdown file
        csv_stats: Optional dict with CSV statistics (mapped, source_gaps, etc.)
    """
    print(f"\nGenerating summary markdown file...")
    
    # Count relationships from Excel analysis
    relationship_counts = df["oscal_relationship"].value_counts().sort_index()
    total = len(df)
    
    # Build markdown content
    lines = [
        "# NIST SP 800-53 Rev 5 to Rev 4 Comparison Summary",
        "",
        "## Overview",
        "",
        f"Total controls analyzed: **{total}**",
        "",
        "## OSCAL Relationship Distribution (Excel Analysis)",
        "",
        "| Relationship | Count | Percentage |",
        "|--------------|-------|------------|"
    ]
    
    for rel, count in relationship_counts.items():
        pct = (count / total * 100) if total > 0 else 0
        lines.append(f"| {rel} | {count} | {pct:.1f}% |")
    
    # Add CSV mapping statistics if provided
    if csv_stats:
        lines.extend([
            "",
            "## CSV Mapping Statistics",
            "",
            f"- **Mapped controls**: {csv_stats['mapped']} (controls with active Rev5↔Rev4 relationships)",
            f"- **Source gaps**: {csv_stats['source_gaps']} (new or restored in Rev5)",
            f"  - New controls (no-relationship): {csv_stats['new_controls']}",
            f"  - Restored controls (restored5): {csv_stats['restored_controls']}",
            f"- **Excluded**: {csv_stats['excluded']} (withdrawn/withdrawn5 controls not in CSV)",
            f"- **Total CSV rows**: {csv_stats['total_rows']}",
        ])
    
    lines.extend([
        "",
        "## Relationship Definitions",
        "",
        "- **equal-to**: No changes at all between Rev 5 and Rev 4",
        "- **equivalent-to**: Cosmetic or discussion-only changes; same substance",
        "- **superset-of**: Rev 5 added requirements (Rev 5 ⊃ Rev 4)",
        "- **subset-of**: Rev 5 removed requirements (Rev 5 ⊂ Rev 4)",
        "- **intersects-with**: Overlapping changes in both directions",
        "- **no-relationship**: New Rev 5 control; no Rev 4 counterpart",
        "- **withdrawn**: Withdrawn in both Rev 4 and Rev 5",
        "- **withdrawn4**: Withdrawn in Rev 4, does not appear in Rev 5",
        "- **restored5**: Withdrawn in Rev 4, restored in Rev 5",
        "- **withdrawn5**: Withdrawn in Rev 5, did not appear in Rev 4 (or was active in Rev 4)",
        "",
        "## Notes",
        "",
        "- Controls with **restored5** relationship are included in the CSV as source gaps",
        "- Controls with **withdrawn**, **withdrawn4**, and **withdrawn5** relationships are excluded from the CSV/JSON output",
        ""
    ])
    
    # Write to file
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, 'w') as f:
        f.write('\n'.join(lines))
    
    print(f"✓ Created {output_file}")

def main():
    if not os.path.exists(INPUT_FILE):
        print(f"Error: {INPUT_FILE} not found.")
        return
    
    print(f"Reading {INPUT_FILE}...")
    
    # Read the Excel file - the main data is in 'Rev4 Rev5 Compared' sheet
    df = pd.read_excel(INPUT_FILE, sheet_name='Rev4 Rev5 Compared', header=0)
    
    # Skip the sub-header row (row 1 in the data)
    df = df.iloc[1:].reset_index(drop=True)
    
    # Standardize column names for easier access
    df.columns = [
        "rev5_id", "rev5_title", "privacy", "low", "med", "high",
        "significant_change", "changed_elements", "change_details",
        "sort_as", "rev4_info",
    ]
    
    # Add OSCAL relationship classification
    print("Classifying OSCAL relationships...")
    df["oscal_relationship"] = df.apply(
        lambda row: classify_relationship(row["changed_elements"], row["change_details"]),
        axis=1
    )
    
    print("\nOSCAL relationship distribution:")
    print(df["oscal_relationship"].value_counts().to_string())
    
    # Generate the relationships Excel file
    generate_relationships_excel(df, INPUT_FILE, RELATIONSHIPS_FILE)
    
    # Now generate the CSV crosswalk
    print(f"\nGenerating CSV crosswalk...")
    
    # Get the relevant columns for CSV generation
    rev5_col = 'rev5_id'
    rev4_col = 'sort_as'
    
    # Check for any withdrawn-error cases
    error_cases = df[df['oscal_relationship'] == 'withdrawn-error']
    if len(error_cases) > 0:
        print(f"\n⚠️  WARNING: Found {len(error_cases)} controls with unexpected withdrawn combinations:")
        print(error_cases[['rev5_id', 'changed_elements', 'change_details']].to_string())
        print("\nThese need manual review!")
    
    # Separate controls into mapped and gap categories
    # Mapped: controls that have active relationships between Rev5 and Rev4
    # Exclude all withdrawn relationships and no-relationship
    # Note: restored5 is handled as a source gap (was withdrawn in Rev4, now restored in Rev5)
    df_mapped = df[
        (df[rev5_col].notna()) &
        (df[rev4_col].notna()) &
        (df['oscal_relationship'] != 'no-relationship') &
        (df['oscal_relationship'] != 'withdrawn') &
        (df['oscal_relationship'] != 'withdrawn4') &
        (df['oscal_relationship'] != 'restored5') &
        (df['oscal_relationship'] != 'withdrawn5') &
        (df['oscal_relationship'] != 'withdrawn-error')
    ].copy()
    
    # Source gaps: Rev5 controls with no Rev4 counterpart
    # Includes: no-relationship (new controls) and restored5 (withdrawn in Rev4, restored in Rev5)
    df_source_gaps = df[
        (df[rev5_col].notna()) &
        ((df['oscal_relationship'] == 'no-relationship') |
         (df['oscal_relationship'] == 'restored5'))
    ].copy()
    
    # Target gaps: None for NIST (withdrawn controls are excluded entirely)
    # We don't include withdrawn, withdrawn4, withdrawn5, or withdrawn-error controls in the CSV/JSON
    df_target_gaps = pd.DataFrame(columns=[rev5_col, rev4_col, 'oscal_relationship'])
    
    print(f"Found {len(df_mapped)} control mappings")
    print(f"Found {len(df_source_gaps)} source gaps (new in Rev5 or restored from Rev4)")
    print(f"Note: Withdrawn controls (withdrawn, withdrawn4, withdrawn5) are excluded from CSV/JSON output")
    
    # Prepare CSV statistics for summary
    csv_stats = {
        'mapped': len(df_mapped),
        'source_gaps': len(df_source_gaps),
        'new_controls': len(df_source_gaps[df_source_gaps['oscal_relationship'] == 'no-relationship']),
        'restored_controls': len(df_source_gaps[df_source_gaps['oscal_relationship'] == 'restored5']),
        'excluded': len(df[(df['oscal_relationship'] == 'withdrawn') |
                          (df['oscal_relationship'] == 'withdrawn4') |
                          (df['oscal_relationship'] == 'withdrawn5')]),
        'total_rows': len(df_mapped) + len(df_source_gaps)
    }
    
    # Generate the summary markdown file with CSV stats
    generate_summary_markdown(df, SUMMARY_FILE, csv_stats)
    
    # Transform the control IDs to OSCAL format for mapped controls
    df_mapped['rev5_oscal'] = df_mapped[rev5_col].apply(transform_rev5_id)
    df_mapped['rev4_oscal'] = df_mapped[rev4_col].apply(transform_rev4_id)
    
    # Build the data rows for mapped controls
    mapped_rows = pd.DataFrame({
        '$$Source_Resource': ["catalogs/NIST_SP-800-53_rev5/catalog.json"] * len(df_mapped),
        '$$Target_Resource': ["catalogs/NIST_SP-800-53_rev4/catalog.json"] * len(df_mapped),
        '$$Map_Source_ID_Ref_list': df_mapped['rev5_oscal'].values,
        '$$Map_Target_ID_Ref_list': df_mapped['rev4_oscal'].values,
        '$$Map_Relationship': df_mapped['oscal_relationship'].values,
        '$Map_Confidence_Score': ["100%"] * len(df_mapped),
        '$Map_Coverage': [""] * len(df_mapped)
    })
    
    # Build rows for source gaps (new Rev5 controls with no Rev4 counterpart)
    if len(df_source_gaps) > 0:
        df_source_gaps['rev5_oscal'] = df_source_gaps[rev5_col].apply(transform_rev5_id)
        # Gap rows must have empty relationship field per OSCAL spec
        # The restored5 classification is tracked internally but not in the OSCAL mapping
        source_gap_rows = pd.DataFrame({
            '$$Source_Resource': ["catalogs/NIST_SP-800-53_rev5/catalog.json"] * len(df_source_gaps),
            '$$Target_Resource': ["catalogs/NIST_SP-800-53_rev4/catalog.json"] * len(df_source_gaps),
            '$$Map_Source_ID_Ref_list': df_source_gaps['rev5_oscal'].values,
            '$$Map_Target_ID_Ref_list': [""] * len(df_source_gaps),  # Empty target = source gap
            '$$Map_Relationship': [""] * len(df_source_gaps),  # Empty relationship for gaps
            '$Map_Confidence_Score': [""] * len(df_source_gaps),
            '$Map_Coverage': [""] * len(df_source_gaps)
        })
    else:
        source_gap_rows = pd.DataFrame(columns=column_names)
    
    # No target gaps - withdrawn controls are excluded
    target_gap_rows = pd.DataFrame(columns=column_names)
    
    # Combine all rows: mapped + source gaps (no target gaps for withdrawn)
    data_rows = pd.concat([mapped_rows, source_gap_rows], ignore_index=True)
    
    # Ensure the columns follow the template order exactly
    data_rows = data_rows[column_names]
    
    # Create the final output including the description row
    # Row 1: Column Names (handled by to_csv header)
    # Row 2: Column Descriptions
    final_df = pd.DataFrame([column_descriptions], columns=column_names)
    final_df = pd.concat([final_df, data_rows], ignore_index=True)
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    # Save to CSV
    final_df.to_csv(OUTPUT_FILE, index=False)
    
    print(f"\n✓ Created {OUTPUT_FILE}")
    print(f"  - Mapped controls: {len(mapped_rows)}")
    print(f"  - Source gaps (new/restored in Rev5): {len(source_gap_rows)}")
    print(f"  - Total rows: {len(data_rows)}")
    print(f"  - Excluded: withdrawn and withdrawn5 controls (not in CSV/JSON)")
    
    # Show a few sample mappings
    print("\nSample mapped controls:")
    for i in range(min(3, len(mapped_rows))):
        source = mapped_rows.iloc[i]['$$Map_Source_ID_Ref_list']
        target = mapped_rows.iloc[i]['$$Map_Target_ID_Ref_list']
        rel = mapped_rows.iloc[i]['$$Map_Relationship']
        print(f"  {source} -> {target} ({rel})")
    
    if len(source_gap_rows) > 0:
        print(f"\nSample source gaps (new/restored in Rev5):")
        for i in range(min(3, len(source_gap_rows))):
            source = source_gap_rows.iloc[i]['$$Map_Source_ID_Ref_list']
            # Check if this is a restored control by looking up in df_source_gaps
            matching_rows = df_source_gaps[df_source_gaps['rev5_oscal'] == source]
            is_restored = False
            if len(matching_rows) > 0:
                is_restored = matching_rows.iloc[0]['oscal_relationship'] == 'restored5'
            label = " (restored from Rev4)" if is_restored else " (new control)"
            print(f"  {source}{label}")
    
    print("\n" + "="*70)
    print("NIST mapping generation complete!")
    print("="*70)
    print(f"\nGenerated files:")
    print(f"  1. {RELATIONSHIPS_FILE}")
    print(f"  2. {SUMMARY_FILE}")
    print(f"  3. {OUTPUT_FILE}")
    print(f"\nNext step: Run 'make nist-json' to generate OSCAL JSON mapping collection")

if __name__ == "__main__":
    main()

# Made with Bob
