"""
add_oscal_relationship.py

Adds an OSCAL relationship column (Rev5 → Rev4) to the NIST SP 800-53
Rev4-to-Rev5 comparison workbook.

Usage:
    python add_oscal_relationship.py <input.xlsx> <output.xlsx>

Dependencies:
    pip install pandas openpyxl
"""

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Classification logic
# ---------------------------------------------------------------------------

ADDS            = ["adds control text", "adds parameter"]
REMOVES         = ["removes parameter", "removes control text"]
CHANGES_CONTROL = ["changes control text", "changes parameter"]
NEUTRAL         = ["changes discussion", "adds discussion", "changes title", "adds to", "n"]
NEW             = ["new base control", "new control enhancement"]


def classify(row: pd.Series) -> str:
    """
    Return an OSCAL relationship label for a single control row.

    Direction: source = Rev5, target = Rev4.

    Labels:
        equal-to        — no changes at all
        equivalent-to   — cosmetic / discussion-only changes; same substance
        superset-of     — Rev5 added requirements  (Rev5 ⊃ Rev4)
        subset-of       — Rev5 removed requirements (Rev5 ⊂ Rev4)
        intersects-with — overlapping changes in both directions
        no-relationship — new Rev5 control; no Rev4 counterpart
        withdrawn       — withdrawn in both Rev4 and Rev5
        withdrawn4      — withdrawn in Rev4, restored in Rev5
        withdrawn5      — active in Rev4, withdrawn in Rev5
    """
    ce = str(row["changed_elements"]).strip().lower() if pd.notna(row["changed_elements"]) else ""
    cd = str(row["change_details"]).strip().lower()   if pd.notna(row["change_details"])   else ""

    # withdrawn4: withdrawn in Rev4 but explicitly restored in Rev5
    if "withdrawn in rev4" in cd and "restored in rev5" in cd:
        return "withdrawn4"

    # withdrawn (both): previously withdrawn in Rev4, also gone in Rev5
    if "previously withdrawn in rev4" in cd:
        return "withdrawn"

    # withdrawn5: active in Rev4, withdrawn in Rev5
    if ce == "withdrawn":
        return "withdrawn5"

    # New controls introduced in Rev5 with no Rev4 counterpart
    if any(n in ce for n in NEW):
        return "no-relationship"

    # Explicitly unchanged
    if ce == "n":
        return "equal-to"

    # Strip neutral-only lines to check for substantive changes
    lines = [l.strip().lower() for l in ce.split("\n") if l.strip()]
    substantive = [l for l in lines if not any(l.startswith(n) for n in NEUTRAL)]

    if not substantive:
        return "equivalent-to"

    has_adds           = any(a in ce for a in ADDS)
    has_removes        = any(r in ce for r in REMOVES)
    has_changes_control = any(c in ce for c in CHANGES_CONTROL)

    if has_changes_control:
        return "intersects-with"
    if has_adds and has_removes:
        return "intersects-with"
    if has_adds and not has_removes:
        return "superset-of"     # Rev5 gained requirements
    if has_removes and not has_adds:
        return "subset-of"       # Rev5 lost requirements

    return "intersects-with"


# ---------------------------------------------------------------------------
# Cell colours for each relationship type
# ---------------------------------------------------------------------------

COLORS = {
    "equal-to":        "C6EFCE",   # green
    "equivalent-to":   "FFEB9C",   # yellow
    "subset-of":       "BDD7EE",   # blue
    "superset-of":     "FCE4D6",   # orange
    "intersects-with": "E2EFDA",   # light green
    "no-relationship": "F2DCDB",   # pink/red
    "withdrawn":       "D9D9D9",   # grey
    "withdrawn4":      "E2CFDD",   # purple-ish
    "withdrawn5":      "C9C9C9",   # dark grey
}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(input_path: str, output_path: str) -> None:
    # ------------------------------------------------------------------
    # 1. Read data with pandas and classify every row
    # ------------------------------------------------------------------
    df = pd.read_excel(input_path, sheet_name="Rev4 Rev5 Compared", header=0)
    df = df.iloc[1:].reset_index(drop=True)   # drop sub-header row
    df.columns = [
        "rev5_id", "rev5_title", "privacy", "low", "med", "high",
        "significant_change", "changed_elements", "change_details",
        "sort_as", "rev4_info",
    ]

    df["oscal_relationship"] = df.apply(classify, axis=1)

    print("OSCAL relationship distribution:")
    print(df["oscal_relationship"].value_counts().to_string())

    # ------------------------------------------------------------------
    # 2. Load the workbook with openpyxl (preserves existing formatting)
    # ------------------------------------------------------------------
    wb = load_workbook(input_path)
    ws = wb["Rev4 Rev5 Compared"]

    new_col = ws.max_column + 2   # leave one blank column as a gap

    # Header row (row 1)
    header = ws.cell(row=1, column=new_col, value="OSCAL Relationship\n(Rev5 → Rev4)")
    header.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header.fill      = PatternFill("solid", fgColor="4472C4")
    header.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Sub-header row (row 2)
    sub = ws.cell(row=2, column=new_col, value="OSCAL Mapping (Rev5→Rev4)")
    sub.font      = Font(bold=True, name="Arial", size=9)
    sub.fill      = PatternFill("solid", fgColor="8EA9C1")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows start at Excel row 3
    for i, row in df.iterrows():
        excel_row    = i + 3
        relationship = row["oscal_relationship"]
        cell         = ws.cell(row=excel_row, column=new_col, value=relationship)
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = PatternFill("solid", fgColor=COLORS.get(relationship, "FFFFFF"))

    ws.column_dimensions[get_column_letter(new_col)].width = 22

    # ------------------------------------------------------------------
    # 3. Save
    # ------------------------------------------------------------------
    wb.save(output_path)
    print(f"\nSaved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python add_oscal_relationship.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
